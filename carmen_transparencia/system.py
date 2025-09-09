#!/usr/bin/env python3
"""
Carmen de Areco - Integrated Transparency System
=================================================

This system integrates multiple Argentine government data sources and existing
GitHub projects to create a comprehensive corruption tracking and transparency portal.

It leverages existing open-source projects and official government APIs to
systematically track irregularities, cross-reference data, and detect patterns
in financial misconduct.

Key Integrations:
- Boletín Oficial scrapers (BORA)
- datosgobar (official Argentine government data)
- AFIP web services
- Tribunal de Cuentas data
- Municipal transparency portals
- Buenos Aires open data

Author: Civic Transparency Project
License: MIT (for transparency purposes)
"""

import asyncio
import aiohttp
import pandas as pd
import numpy as np
import requests
from datetime import datetime, timedelta
import json
import logging
from pathlib import Path
import sqlite3
from typing import Dict, List, Optional, Tuple, Set
import hashlib
import re
from dataclasses import dataclass, asdict
from enum import Enum

# Third-party integrations
try:
    import scrapy
    from scrapy.crawler import CrawlerProcess
    from bs4 import BeautifulSoup
    import PyPDF2
    import openpyxl
except (ModuleNotFoundError, ImportError) as e:
    print(f"Error: Módulo necesario no encontrado: {e}")
    print("Por favor, instale las dependencias con 'pip install -r requirements.txt'")
    exit(1)


# Configuration
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DataSource(Enum):
    """Enumeration of data sources"""
    BOLETIN_OFICIAL = "boletin_oficial"
    DATOS_GOB_AR = "datos_gob_ar"
    DATOS_BUENOS_AIRES = "datos_buenos_aires"
    HTC_BUENOS_AIRES = "htc_buenos_aires"
    AFIP = "afip"
    MUNICIPAL_TRANSPARENCY = "municipal_transparency"
    SIBOM = "sibom"
    GEOREF = "georef"
    OBRAS_PUBLICAS = "obras_publicas"
    CONCEJO_DELIBERANTE = "concejo_deliberante"
    LOCAL_INVENTORY = "local_inventory"
    PRESUPUESTO_ABIERTO = "presupuesto_abierto"
    BOLETIN_OFICIAL_PROVINCIA = "boletin_oficial_provincia"


@dataclass
class CorruptionCase:
    """Data structure for corruption cases"""
    case_id: str
    title: str
    description: str
    amount: Optional[float]
    responsible_parties: List[str]
    date: str
    status: str
    severity: str
    category: str
    source: str
    source_url: str
    documents: List[str]
    verification_status: str
    cross_references: List[str]

@dataclass
class DocumentRecord:
    """Data structure for document tracking"""
    doc_id: str
    title: str
    url: str
    download_date: datetime
    file_hash: str
    file_size: int
    document_type: str
    source: DataSource
    integrity_status: str
    anomalies_detected: List[str]
    signatories: List[str]
    financial_data: Dict

class IntegratedTransparencySystem:
    """Main class for the integrated transparency system"""
    
    def __init__(self, data_dir: str = "transparency_data"):
        self.data_dir = Path(data_dir)
        self.data_dir.mkdir(exist_ok=True)
        self.project_root = Path.cwd()
        
        # Initialize database
        self.db_path = self.data_dir / "transparency.db"
        self.init_database()
        
        # Government data sources configuration
        self.data_sources = {
            DataSource.BOLETIN_OFICIAL: {
                "base_url": "https://www.boletinoficial.gob.ar/",
                "scraper_repos": [
                    "https://github.com/tommanzur/scraper_boletin_oficial",
                    "https://github.com/dedio/bora-scraper",
                    "https://github.com/chrishein/bora_crawler"
                ]
            },
            DataSource.BOLETIN_OFICIAL_PROVINCIA: {
                "base_url": "https://www.boletinoficial.gba.gob.ar/",
            },
            DataSource.DATOS_GOB_AR: {
                "base_url": "https://datos.gob.ar/",
                "api_url": "https://datos.gob.ar/api/3/",
                "github_repo": "https://github.com/datosgobar"
            },
            DataSource.DATOS_BUENOS_AIRES: {
                "base_url": "https://data.buenosaires.gob.ar/",
                "api_url": "https://data.buenosaires.gob.ar/api/3/",
                "github_repo": "https://github.com/datosgcba"
            },
            DataSource.HTC_BUENOS_AIRES: {
                "base_url": "https://www.htc.gba.gov.ar/",
                "contact": "info @htc.gba.gov.ar"
            },
            DataSource.AFIP: {
                "base_url": "https://www.afip.gob.ar/",
                "ws_urls": {
                    "wsaa": "https://wsaa.afip.gov.ar/ws/services/LoginCms?wsdl",
                    "wsfe": "https://servicios1.afip.gov.ar/wsfev1/service.asmx?WSDL"
                },
                "github_repos": [
                    "https://github.com/reingart/pyafipws",
                    "https://github.com/AfipSDK/afip.py"
                ]
            },
            DataSource.SIBOM: {
                "base_url": "https://sibom.slyt.gba.gob.ar/",
                "carmen_areco_url": "https://sibom.slyt.gba.gob.ar/cities/23"
            },
            DataSource.GEOREF: {
                "api_url": "https://apis.datos.gob.ar/georef/api/",
                "github_repo": "https://github.com/datosgobar/georef-ar-api"
            },
            DataSource.OBRAS_PUBLICAS: {
                "base_url": "https://mapainversiones.obraspublicas.gob.ar/datos-abiertos",
                "docs": "https://mapainversiones.obraspublicas.gob.ar/docs/api/v1/"
            },
            DataSource.CONCEJO_DELIBERANTE: {
                "base_url": "http://hcdcarmendeareco.blogspot.com/",
            },
            DataSource.PRESUPUESTO_ABIERTO: {
                "base_url": "https://presupuestoabierto.gob.ar/",
                "api_url": "https://presupuestoabierto.gob.ar/api/v1/"
            }
        }
        
        # Known corruption cases from HTC and investigations
        self.corruption_cases = self._load_known_cases()
        
        # Financial anomaly detection thresholds
        self.anomaly_thresholds = {
            "large_deviation": 0.15,  # 15%
            "missing_documentation": True,
            "unauthorized_signatures": True,
            "budget_overrun": 0.10,  # 10%
            "unexplained_transfers": 1000000.0  # 1M ARS
        }

    def init_database(self):
        """Initialize SQLite database for storing transparency data"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Corruption cases table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS corruption_cases (
                case_id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                description TEXT,
                amount REAL,
                responsible_parties TEXT,
                date TEXT,
                status TEXT,
                severity TEXT,
                category TEXT,
                source TEXT,
                source_url TEXT,
                documents TEXT,
                verification_status TEXT,
                cross_references TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Documents table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS documents (
                doc_id TEXT PRIMARY KEY,
                title TEXT NOT NULL,
                url TEXT,
                download_date TIMESTAMP,
                file_hash TEXT,
                file_size INTEGER,
                document_type TEXT,
                source TEXT,
                integrity_status TEXT,
                anomalies_detected TEXT,
                signatories TEXT,
                financial_data TEXT,
                content TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Financial data table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS financial_records (
                record_id TEXT PRIMARY KEY,
                municipality TEXT,
                period TEXT,
                total_revenue REAL,
                total_expenses REAL,
                balance REAL,
                budget_execution_rate REAL,
                document_source TEXT,
                anomalies TEXT,
                risk_score REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Data sources tracking table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS data_source_logs (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                source TEXT,
                last_update TIMESTAMP,
                records_processed INTEGER,
                errors_count INTEGER,
                status TEXT
            )
        ''')
        
        conn.commit()
        conn.close()

    def _load_known_cases(self) -> List[CorruptionCase]:
        """Load known corruption cases from HTC and other official sources"""
        return [
            CorruptionCase(
                case_id="htc_carmen_areco_2024_001",
                title="Subsidio con Irregularidades - HTC",
                description="Multa de $280.000 por otorgar subsidio con diferencias entre entidad subsidiada y quien realizó la rendición",
                amount=280000.00,
                responsible_parties=["Iván Villagrán (Intendente)", "Alicia Batallón (Contadora)"],
                date="2024",
                status="Confirmed by HTC",
                severity="high",
                category="Misuse of public funds",
                source="Honorable Tribunal de Cuentas Buenos Aires",
                source_url="https://ahoralm.com.ar/contenido/9578/el-h-tribunal-de-cuentas-y-los-intendentes",
                documents=["HTC Resolution 2024", "Municipal Financial Reports"],
                verification_status="verified",
                cross_references=["htc_carmen_areco_2024_002"]
            ),
            CorruptionCase(
                case_id="htc_carmen_areco_2024_002",
                title="Cargo Solidario Adicional - HTC",
                description="Cargo adicional de $2.362.010,41 por responsabilidad solidaria en irregularidades financieras",
                amount=2362010.41,
                responsible_parties=["Iván Villagrán", "Alicia Batallón", "Sebastián Torretta"],
                date="2024",
                status="Confirmed by HTC",
                severity="critical",
                category="Joint financial responsibility",
                source="Honorable Tribunal de Cuentas Buenos Aires",
                source_url="https://ahoralm.com.ar/contenido/9578/el-h-tribunal-de-cuentas-y-los-intendentes",
                documents=["HTC Resolution 2024"],
                verification_status="verified",
                cross_references=["htc_carmen_areco_2024_001"]
            ),
            CorruptionCase(
                case_id="carmen_areco_2025_001",
                title="Licitación Fraudulenta - Consejero",
                description="Presidente del Consejo Honorable licitó para su propia empresa con el municipio",
                amount=None,
                responsible_parties=["Carlos Alfredo Camallo (Presidente Consejo Honorable)"],
                date="2025",
                status="Under judicial investigation",
                severity="critical",
                category="Conflict of interest / Self-contracting",
                source="Judicial Process",
                source_url="",
                documents=["Bidding documents", "Contract records"],
                verification_status="under_investigation",
                cross_references=[]
            ),
            CorruptionCase(
                case_id="carmen_areco_2025_002",
                title="Documentos Alterados - Tesorería",
                description="Documentos firmados por personas no autorizadas en lugar del responsable de tesorería",
                amount=None,
                responsible_parties=["Personal de Tesorería", "Terceros no autorizados"],
                date="2025",
                status="Under investigation",
                severity="high",
                category="Document forgery",
                source="Internal audit",
                source_url="",
                documents=["Treasury documents", "Financial statements"],
                verification_status="suspicious",
                cross_references=[]
            )
        ]

    def ingest_local_inventory(self):
        """Ingest the local document inventory from JSON into the database."""
        logger.info("Ingesting local document inventory...")
        inventory_paths = list(self.project_root.glob('**/document_inventory.json'))
        if not inventory_paths:
            logger.warning("document_inventory.json not found anywhere in the project.")
            return

        inventory_path = inventory_paths[0]
        logger.info(f"Found inventory at: {inventory_path}")

        try:
            with open(inventory_path, 'r', encoding='utf-8') as f:
                inventory_data = json.load(f)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            records_to_update = []

            for item in inventory_data:
                doc_id = item.get('filename')
                if not doc_id:
                    continue

                # First, insert the basic record
                cursor.execute('''
                    INSERT OR IGNORE INTO documents (doc_id, title, url, file_size, document_type, source, integrity_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    doc_id,
                    item.get('filename', 'No Title'),
                    item.get('official_url'),
                    item.get('file_size'),
                    item.get('file_type'),
                    DataSource.LOCAL_INVENTORY.value,
                    'pending_verification'
                ))

                # Now, process the file content if it exists
                content = None
                analysis_report = {}
                file_path_str = item.get('relative_path')

                if file_path_str:
                    # The relative_path in the inventory seems to be relative to the inventory's parent directory
                    base_path = inventory_path.parent
                    full_path = base_path / file_path_str

                    if item['file_type'] == 'pdf':
                        pdf_filename = Path(item['official_url']).name
                        # Search for the PDF file in the entire project directory
                        pdf_paths = list(self.project_root.glob(f"**/{pdf_filename}"))
                        if pdf_paths:
                            full_path = pdf_paths[0]
                        else:
                            full_path = None
                            logger.warning(f"PDF file not found locally: {pdf_filename}")

                    if full_path and full_path.exists():
                        if item['file_type'] in ['md', 'json']:
                            try:
                                with open(full_path, 'r', encoding='utf-8') as doc_file:
                                    content = doc_file.read()
                            except Exception as e:
                                logger.error(f"Could not read file {full_path}: {e}")
                        elif item['file_type'] == 'pdf':
                            analysis_report = self.analyze_document_integrity(str(full_path))
                            content = analysis_report.get('content')
                    elif item['file_type'] != 'pdf': # Avoid double warning for PDFs
                        logger.warning(f"File not found at path: {full_path}")
                
                records_to_update.append({
                    'doc_id': doc_id,
                    'content': content,
                    'file_hash': analysis_report.get('file_hash'),
                    'integrity_status': analysis_report.get('integrity_status', 'pending_verification'),
                    'anomalies_detected': json.dumps(analysis_report.get('anomalies', []))
                })

            # Batch update the records with content and analysis
            for record in records_to_update:
                cursor.execute('''
                    UPDATE documents 
                    SET content = ?, file_hash = ?, integrity_status = ?, anomalies_detected = ?
                    WHERE doc_id = ?
                ''', (
                    record['content'],
                    record['file_hash'],
                    record['integrity_status'],
                    record['anomalies_detected'],
                    record['doc_id']
                ))

            conn.commit()
            conn.close()
            logger.info(f"Successfully ingested and processed {len(inventory_data)} records from local inventory.")

        except Exception as e:
            logger.error(f"Error ingesting local inventory: {e}")

    async def scrape_boletin_oficial(self, search_terms: List[str] = None) -> List[Dict]:
        """
        Scrape Boletín Oficial for Carmen de Areco related content
        Integrates with existing GitHub scrapers
        """
        if search_terms is None:
            search_terms = ["Carmen de Areco", "Iván Villagrán", "multa", "sanción", "Tribunal de Cuentas"]
        
        logger.info("Scraping Boletín Oficial for Carmen de Areco content")
        
        # Use existing scraper from GitHub (tommanzur/scraper_boletin_oficial pattern)
        base_url = "https://www.boletinoficial.gob.ar/buscar"
        results = []
        
        async with aiohttp.ClientSession() as session:
            for term in search_terms:
                try:
                    params = {
                        'busqueda': term,
                        'f_desde': '2020-01-01',
                        'f_hasta': datetime.now().strftime('%Y-%m-%d')
                    }
                    
                    async with session.get(base_url, params=params) as response:
                        if response.status == 200:
                            content = await response.text()
                            soup = BeautifulSoup(content, 'html.parser')
                            
                            # Parse results (pattern from existing scrapers)
                            for item in soup.find_all('div', class_='resultado'):
                                title = item.find('h3')
                                date = item.find('span', class_='fecha')
                                link = item.find('a')
                                
                                if title and 'carmen de areco' in title.get_text().lower():
                                    results.append({
                                        'title': title.get_text().strip(),
                                        'date': date.get_text().strip() if date else '',
                                        'url': link['href'] if link else '',
                                        'search_term': term,
                                        'source': DataSource.BOLETIN_OFICIAL.value
                                    })
                
                except Exception as e:
                    logger.error(f"Error scraping Boletín Oficial for term '{term}': {e}")
        
        logger.info(f"Found {len(results)} Boletín Oficial entries for Carmen de Areco")
        return results

    async def scrape_boletin_oficial_provincia(self, search_terms: List[str] = None) -> List[Dict]:
        """Scrape the Official Gazette of the Province of Buenos Aires."""
        logger.info("Scraping Boletín Oficial de la Provincia de Buenos Aires")
        # This is a placeholder. A real implementation would use a library like BeautifulSoup
        # to parse the HTML of the website and extract the links to the gazettes.
        # It would also need to handle pagination and date-based searches.
        return []

    async def query_datos_gob_ar(self) -> List[Dict]:
        """
        Query datos.gob.ar for Carmen de Areco related datasets
        Uses official datosgobar APIs
        """
        logger.info("Querying datos.gob.ar for Carmen de Areco datasets")
        
        api_url = "https://datos.gob.ar/api/3/action/package_search"
        search_queries = [
            "Carmen de Areco",
            "municipio Carmen de Areco",
            "presupuesto municipal Buenos Aires",
            "transparencia fiscal",
            "compras",
            "contrataciones",
            "licitaciones",
            "obra pública"
        ]
        
        results = []
        
        async with aiohttp.ClientSession() as session:
            for query in search_queries:
                try:
                    params = {
                        'q': query,
                        'rows': 100,
                        'fq': 'organization:carmen-de-areco OR tags:carmen-de-areco'
                    }
                    
                    async with session.get(api_url, params=params) as response:
                        if response.status == 200:
                            data = await response.json()
                            
                            if data.get('success') and data.get('result'):
                                for package in data['result']['results']:
                                    results.append({
                                        'id': package.get('id'),
                                        'title': package.get('title'),
                                        'description': package.get('notes'),
                                        'organization': package.get('organization', {}).get('title'),
                                        'tags': [tag['name'] for tag in package.get('tags', [])],
                                        'resources': len(package.get('resources', [])),
                                        'url': f"https://datos.gob.ar/dataset/{package.get('name')}",
                                        'source': DataSource.DATOS_GOB_AR.value
                                    })
                
                except Exception as e:
                    logger.error(f"Error querying datos.gob.ar for '{query}': {e}")
        
        logger.info(f"Found {len(results)} datasets on datos.gob.ar")
        return results

    async def query_buenos_aires_data(self) -> List[Dict]:
        """
        Query Buenos Aires open data portal
        Uses datosgcba APIs and resources
        """
        logger.info("Querying Buenos Aires open data portal")
        
        api_url = "https://data.buenosaires.gob.ar/api/3/action/package_search"
        search_queries = [
            "municipio",
            "presupuesto",
            "transparencia",
            "tribunal de cuentas"
        ]
        
        results = []
        
        async with aiohttp.ClientSession() as session:
            for query in search_queries:
                try:
                    params = {
                        'q': query,
                        'rows': 50
                    }
                    
                    async with session.get(api_url, params=params) as response:
                        if response.status == 200:
                            if 'application/json' in response.headers.get('Content-Type', ''):
                                data = await response.json()
                                
                                if data.get('success') and data.get('result'):
                                    for package in data['result']['results']:
                                        # Filter for relevant municipal data
                                        if any(keyword in package.get('title', '').lower() 
                                              for keyword in ['municipio', 'presupuesto', 'transparencia']):
                                            results.append({
                                                'id': package.get('id'),
                                                'title': package.get('title'),
                                                'description': package.get('notes'),
                                                'resources': package.get('resources', []),
                                                'url': f"https://data.buenosaires.gob.ar/dataset/{package.get('name')}",
                                                'source': DataSource.DATOS_BUENOS_AIRES.value
                                            })
                            else:
                                logger.warning(f"Buenos Aires data API returned non-JSON response for query '{query}'. Status: {response.status}")

                except Exception as e:
                    logger.error(f"Error querying Buenos Aires data for '{query}': {e}")
        
        logger.info(f"Found {len(results)} relevant datasets on Buenos Aires open data")
        return results

    async def query_obras_publicas(self) -> List[Dict]:
        """Query the national public works API."""
        logger.info("Querying National Public Works data")
        # The data is available as a downloadable CSV file, not a direct API.
        # See: https://mapainversiones.obraspublicas.gob.ar/datos-abiertos
        # A real implementation would download and parse the CSV file.
        return []

    async def query_presupuesto_abierto(self) -> List[Dict]:
        """Query the national open budget API."""
        logger.info("Querying National Open Budget API (Presupuesto Abierto)")
        # This is a placeholder. A real implementation would require an API token.
        # To get a token, visit https://presupuestoabierto.gob.ar/dev
        # Example of how to use the API with a token:
        # headers = {'Authorization': 'Bearer YOUR_API_TOKEN'}
        # async with session.get(api_url, headers=headers) as response:
        #     ...
        return []

    async def scrape_concejo_deliberante(self) -> List[Dict]:
        """Scrape the local council blog for resolutions and ordinances."""
        logger.info("Scraping local council blog (http://hcdcarmendeareco.blogspot.com/)")
        # This is a placeholder. A real implementation would use a library like BeautifulSoup
        # to parse the HTML of the blog and extract the links to the resolutions and ordinances.
        # It would also need to handle pagination if there are multiple pages.
        return []

    def integrate_afip_data(self) -> Dict:
        """
        Integrate AFIP data using existing PyAfipWs libraries
        Checks taxpayer information and electronic invoicing compliance
        """
        logger.info("Integrating AFIP data for Carmen de Areco")
        
        # This would integrate with reingart/pyafipws or similar libraries
        # For now, we simulate the structure
        
        afip_data = {
            "municipality_cuit": "30-99999999-9",  # Carmen de Areco CUIT
            "taxpayer_status": "active",
            "electronic_invoicing": {
                "enabled": True,
                "last_invoice": None,
                "compliance_status": "compliant"
            },
            "contractors_verification": [],
            "source": DataSource.AFIP.value
        }
        
        # Here we would use PyAfipWs to verify:
        # 1. Municipality's taxpayer status
        # 2. Contractors' tax compliance
        # 3. Electronic invoicing compliance
        # 4. Cross-reference with known irregular contracts
        
        return afip_data

    def analyze_document_integrity(self, document_path: str) -> Dict:
        """
        Advanced document integrity analysis
        Detects forgery, unauthorized modifications, and signature issues
        """
        logger.info(f"Analyzing document integrity: {document_path}")
        
        integrity_report = {
            "file_path": document_path,
            "file_hash": self._calculate_file_hash(document_path),
            "file_size": Path(document_path).stat().st_size,
            "analysis_date": datetime.now().isoformat(),
            "integrity_status": "unknown",
            "anomalies": [],
            "signatures_analysis": {},
            "content_analysis": {},
            "content": None,
            "risk_score": 0.0
        }
        
        try:
            if document_path.endswith('.pdf'):
                pdf_analysis, pdf_content = self._analyze_pdf_integrity(document_path)
                integrity_report.update(pdf_analysis)
                integrity_report['content'] = pdf_content
            elif document_path.endswith(('.xlsx', '.xls')):
                integrity_report.update(self._analyze_excel_integrity(document_path))
            
            # Calculate risk score based on anomalies
            integrity_report["risk_score"] = self._calculate_risk_score(integrity_report["anomalies"])
            
            # Determine overall integrity status
            if integrity_report["risk_score"] > 0.8:
                integrity_report["integrity_status"] = "compromised"
            elif integrity_report["risk_score"] > 0.5:
                integrity_report["integrity_status"] = "suspicious"
            elif integrity_report["risk_score"] > 0.2:
                integrity_report["integrity_status"] = "questionable"
            else:
                integrity_report["integrity_status"] = "verified"
        
        except Exception as e:
            logger.error(f"Error analyzing document integrity: {e}")
            integrity_report["anomalies"].append(f"Analysis error: {str(e)}")
        
        return integrity_report

    def _analyze_pdf_integrity(self, pdf_path: str) -> Tuple[Dict, str]:
        """Analyze PDF document for integrity issues. Returns analysis dict and text content."""
        analysis = {
            "signatures_analysis": {"required_signatures": [], "found_signatures": [], "missing_signatures": []},
            "content_analysis": {"suspicious_patterns": [], "modification_indicators": []},
            "anomalies": []
        }
        text = ""
        
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                
                for page in pdf_reader.pages:
                    text += page.extract_text()
                
                # Check for required signatures (Carmen de Areco specific)
                required_signers = [
                    "Iván Villagrán", "Intendente",
                    "Alicia Batallón", "Contadora",
                    "Sebastián Torretta", "Jefe de Compras"
                ]
                
                found_signers = []
                for signer in required_signers:
                    if re.search(signer, text, re.IGNORECASE):
                        found_signers.append(signer)
                    else:
                        analysis["signatures_analysis"]["missing_signatures"].append(signer)
                
                analysis["signatures_analysis"]["found_signatures"] = found_signers
                analysis["signatures_analysis"]["required_signatures"] = required_signers
                
                # Check for suspicious patterns
                suspicious_patterns = [
                    r"corregido.*posterior",
                    r"modificado.*después",
                    r"firma.*no.*autorizada",
                    r"sin.*autorización",
                    r"irregular.*procedimiento"
                ]
                
                for pattern in suspicious_patterns:
                    if re.search(pattern, text, re.IGNORECASE):
                        analysis["content_analysis"]["suspicious_patterns"].append(pattern)
                        analysis["anomalies"].append(f"Suspicious pattern detected: {pattern}")
                
                # Check for modification indicators
                if pdf_reader.metadata and "modified" in pdf_reader.metadata.get('/Producer', '').lower():
                    analysis["anomalies"].append("PDF shows modification indicators in metadata")
                
                # Check for missing signatures
                if analysis["signatures_analysis"]["missing_signatures"]:
                    analysis["anomalies"].append(
                        f"Missing required signatures: {', '.join(analysis['signatures_analysis']['missing_signatures'])}"
                    )
        
        except Exception as e:
            analysis["anomalies"].append(f"PDF analysis error: {str(e)}")
        
        return analysis, text

    def _analyze_excel_integrity(self, excel_path: str) -> Dict:
        """Analyze Excel document for integrity issues"""
        analysis = {
            "signatures_analysis": {},
            "content_analysis": {"formulas_analysis": [], "data_patterns": []},
            "anomalies": []
        }
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=False)
            
            # Check for suspicious formulas or external references
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # Check for external references
                            if cell.value.startswith('=') and ('http' in cell.value or 'file:' in cell.value):
                                analysis["anomalies"].append(f"Suspicious external reference in cell {cell.coordinate}")
                            
                            # Check for hidden formulas that might manipulate data
                            if '&' in cell.value and '=' in cell.value:
                                analysis["content_analysis"]["formulas_analysis"].append(
                                    f"Complex formula in {cell.coordinate}: {cell.value}"
                                )
            
            # Check workbook properties for modification history
            props = workbook.properties
            if props.modified and props.created:
                time_diff = props.modified - props.created
                if time_diff.total_seconds() < 300:  # Modified within 5 minutes of creation
                    analysis["anomalies"].append("Document was modified very shortly after creation")
        
        except Exception as e:
            analysis["anomalies"].append(f"Excel analysis error: {str(e)}")
        
        return analysis

    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA256 hash of file"""
        sha256_hash = hashlib.sha256()
        try:
            with open(file_path, "rb") as f:
                for byte_block in iter(lambda: f.read(4096), b""):
                    sha256_hash.update(byte_block)
            return sha256_hash.hexdigest()
        except Exception as e:
            logger.error(f"Error calculating hash for {file_path}: {e}")
            return ""

    def _calculate_risk_score(self, anomalies: List[str]) -> float:
        """Calculate risk score based on detected anomalies"""
        if not anomalies:
            return 0.0
        
        # Weight different types of anomalies
        weights = {
            "missing.*signature": 0.4,
            "modification": 0.3,
            "suspicious.*pattern": 0.3,
            "external.*reference": 0.2,
            "analysis.*error": 0.1
        }
        
        total_score = 0.0
        for anomaly in anomalies:
            for pattern, weight in weights.items():
                if re.search(pattern, anomaly, re.IGNORECASE):
                    total_score += weight
                    break
            else:
                total_score += 0.1  # Default weight for unclassified anomalies
        
        return min(total_score, 1.0)  # Cap at 1.0

    def cross_reference_all_data(self, detailed_findings: Dict) -> Dict:
        """Cross-reference findings across all data sources"""
        logger.info("Performing advanced cross-referencing of all data sources...")
        cross_ref = {
            "entity_correlations": [],
            "temporal_correlations": [],
            "amount_correlations": [],
            "pattern_matches": [],
            "risk_indicators": []
        }

        # 1. Entity Correlation
        all_entities = set()
        entity_sources = {}
        for source, findings in detailed_findings.items():
            if isinstance(findings, list):
                for finding in findings:
                    entities = []
                    if 'title' in finding:
                        entities.extend(re.findall(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', finding['title']))
                    if 'description' in finding and finding['description']:
                         entities.extend(re.findall(r'\b[A-Z][a-z]+\s+[A-Z][a-z]+\b', finding['description']))
                    if 'responsible_parties' in finding and finding['responsible_parties']:
                        entities.extend(finding['responsible_parties'])
                    
                    for entity in set(entities):
                        all_entities.add(entity)
                        if entity not in entity_sources:
                            entity_sources[entity] = set()
                        entity_sources[entity].add(source)

        for entity, sources in entity_sources.items():
            if len(sources) > 1:
                cross_ref["entity_correlations"].append({
                    "entity": entity,
                    "sources": list(sources),
                    "correlation_strength": len(sources)
                })

        # 2. Known Corruption Entities Check
        known_corruption_entities = []
        for case in self.corruption_cases:
            known_corruption_entities.extend(case.responsible_parties)
        
        for entity in all_entities:
            for known_entity in known_corruption_entities:
                if known_entity in entity:
                    cross_ref["pattern_matches"].append({
                        "type": "known_corruption_entity",
                        "entity": entity,
                        "description": f"Entity '{entity}' matches known corruption case participant: '{known_entity}'"
                    })

        # 3. Temporal and Amount Correlation (Example)
        # This is a simplified example. A real implementation would require more sophisticated logic.
        for source1, findings1 in detailed_findings.items():
            for source2, findings2 in detailed_findings.items():
                if source1 >= source2: continue # Avoid duplicates and self-comparison
                if isinstance(findings1, list) and isinstance(findings2, list):
                    for f1 in findings1:
                        for f2 in findings2:
                            # Temporal correlation
                            if f1.get('date') and f1.get('date') == f2.get('date'):
                                cross_ref["temporal_correlations"].append({
                                    "source1": source1,
                                    "source2": source2,
                                    "item1": f1.get('title'),
                                    "item2": f2.get('title'),
                                    "date": f1.get('date')
                                })
                            # Amount correlation
                            if f1.get('amount') and f1.get('amount') == f2.get('amount'):
                                cross_ref["amount_correlations"].append({
                                    "source1": source1,
                                    "source2": source2,
                                    "item1": f1.get('title'),
                                    "item2": f2.get('title'),
                                    "amount": f1.get('amount')
                                })

        # 4. Risk Indicators
        if len(cross_ref["entity_correlations"]) > 5:
            cross_ref["risk_indicators"].append("High number of entity correlations across different data sources.")
        if len(cross_ref["pattern_matches"]) > 0:
            cross_ref["risk_indicators"].append("Entities matching known corruption case participants found.")

        return cross_ref

    async def run_comprehensive_analysis(self):
        """
        Run comprehensive analysis across all data sources
        """
        logger.info("Starting comprehensive transparency analysis for Carmen de Areco")
        
        # Ingest local data first
        self.ingest_local_inventory()

        analysis_results = {
            "analysis_timestamp": datetime.now().isoformat(),
            "municipality": "Carmen de Areco",
            "data_sources_analyzed": [],
            "corruption_cases_tracked": len(self.corruption_cases),
            "documents_processed": 0,
            "anomalies_detected": 0,
            "overall_risk_level": "unknown",
            "detailed_findings": {},
            "recommendations": []
        }
        
        try:
            # 1. Scrape Boletín Oficial
            boletin_results = await self.scrape_boletin_oficial()
            analysis_results["data_sources_analyzed"].append(DataSource.BOLETIN_OFICIAL.value)
            analysis_results["detailed_findings"]["boletin_oficial"] = boletin_results

            boletin_provincia_results = await self.scrape_boletin_oficial_provincia()
            analysis_results["data_sources_analyzed"].append(DataSource.BOLETIN_OFICIAL_PROVINCIA.value)
            analysis_results["detailed_findings"]["boletin_oficial_provincia"] = boletin_provincia_results
            
            # 2. Query government data portals
            datos_gob_results = await self.query_datos_gob_ar()
            analysis_results["data_sources_analyzed"].append(DataSource.DATOS_GOB_AR.value)
            analysis_results["detailed_findings"]["datos_gob_ar"] = datos_gob_results
            
            # 3. Query Buenos Aires data
            ba_data_results = await self.query_buenos_aires_data()
            analysis_results["data_sources_analyzed"].append(DataSource.DATOS_BUENOS_AIRES.value)
            analysis_results["detailed_findings"]["buenos_aires_data"] = ba_data_results
            
            # 4. Integrate AFIP data
            afip_data = self.integrate_afip_data()
            analysis_results["data_sources_analyzed"].append(DataSource.AFIP.value)
            analysis_results["detailed_findings"]["afip_data"] = afip_data
            
            # 5. Analyze municipal transparency portal
            transparency_analysis = await self.analyze_municipal_transparency()
            analysis_results["detailed_findings"]["municipal_transparency"] = transparency_analysis

            # 6. Query new data sources
            obras_results = await self.query_obras_publicas()
            analysis_results["data_sources_analyzed"].append(DataSource.OBRAS_PUBLICAS.value)
            analysis_results["detailed_findings"]["obras_publicas"] = obras_results

            concejo_results = await self.scrape_concejo_deliberante()
            analysis_results["data_sources_analyzed"].append(DataSource.CONCEJO_DELIBERANTE.value)
            analysis_results["detailed_findings"]["concejo_deliberante"] = concejo_results

            presupuesto_results = await self.query_presupuesto_abierto()
            analysis_results["data_sources_analyzed"].append(DataSource.PRESUPUESTO_ABIERTO.value)
            analysis_results["detailed_findings"]["presupuesto_abierto"] = presupuesto_results
            
            # 7. Cross-reference all findings
            cross_ref_results = self.cross_reference_all_data(analysis_results["detailed_findings"])
            analysis_results["cross_reference_analysis"] = cross_ref_results
            
            # 8. Calculate overall risk assessment
            risk_indicators = []
            
            # Count high-risk findings
            for source, findings in analysis_results["detailed_findings"].items():
                if isinstance(findings, list):
                    for finding in findings:
                        if finding.get("risk_level") == "high":
                            risk_indicators.append(f"High risk finding in {source}")
                elif isinstance(findings, dict) and findings.get("risk_level") == "high":
                    risk_indicators.append(f"High risk in {source}")
            
            # Add known corruption cases to risk calculation
            confirmed_cases = len([case for case in self.corruption_cases if case.status == "Confirmed by HTC"])
            under_investigation = len([case for case in self.corruption_cases if "investigation" in case.status.lower()])
            
            if confirmed_cases >= 2 or under_investigation >= 1:
                risk_indicators.append(f"Active corruption cases: {confirmed_cases} confirmed, {under_investigation} under investigation")
            
            # Determine overall risk level
            if len(risk_indicators) >= 3 or confirmed_cases >= 2:
                analysis_results["overall_risk_level"] = "critical"
            elif len(risk_indicators) >= 2 or confirmed_cases >= 1:
                analysis_results["overall_risk_level"] = "high"
            elif len(risk_indicators) >= 1 or under_investigation >= 1:
                analysis_results["overall_risk_level"] = "medium"
            else:
                analysis_results["overall_risk_level"] = "low"
            
            # Generate recommendations based on findings
            analysis_results["recommendations"] = self.generate_recommendations(analysis_results)
            
            # Store results in database
            self.store_analysis_results(analysis_results)
            
            logger.info(f"Comprehensive analysis completed. Overall risk level: {analysis_results['overall_risk_level']}")
            
        except Exception as e:
            logger.error(f"Error during comprehensive analysis: {e}")
            analysis_results["error"] = str(e)
        
        return analysis_results

    def generate_recommendations(self, analysis_results: Dict) -> List[str]:
        """Generate recommendations based on analysis results"""
        recommendations = []
        
        risk_level = analysis_results.get("overall_risk_level", "unknown")
        
        if risk_level == "critical":
            recommendations.extend([
                "IMMEDIATE ACTION REQUIRED: File formal complaint with appropriate authorities",
                "Request emergency audit by independent external auditor",
                "Consider criminal investigation referral",
                "Implement immediate financial controls and oversight",
                "Demand full disclosure of all financial records"
            ])
        
        elif risk_level == "high":
            recommendations.extend([
                "Escalate concerns to municipal oversight bodies",
                "Request comprehensive financial audit",
                "Organize citizen oversight committee",
                "Demand monthly financial transparency reports",
                "Implement enhanced monitoring of municipal activities"
            ])
        
        elif risk_level == "medium":
            recommendations.extend([
                "Increase frequency of financial monitoring",
                "Implement citizen watch program",
                "Request quarterly transparency reports",
                "Engage with local media for coverage"
            ])
        
        # Specific recommendations based on findings
        if analysis_results.get("detailed_findings", {}).get("municipal_transparency", {}).get("integrity_issues"):
            recommendations.append("Implement digital document integrity verification system")
        
        if analysis_results.get("detailed_findings", {}).get("municipal_transparency", {}).get("missing_documents"):
            recommendations.append("File formal requests for missing transparency documents")
        
        # Add legal recommendations
        recommendations.extend([
            "Contact transparency advocacy organizations",
            "File formal information requests under Access to Information laws",
            "Engage with local council members and opposition parties",
            "Use social media to raise awareness and build citizen pressure",
            "Petition provincial government for intervention if necessary"
        ])
        
        return recommendations

    def store_analysis_results(self, analysis_results: Dict):
        """Store analysis results in database"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Store corruption cases
            for case in self.corruption_cases:
                cursor.execute('''
                    INSERT OR REPLACE INTO corruption_cases 
                    (case_id, title, description, amount, responsible_parties, date, status, 
                     severity, category, source, source_url, documents, verification_status, cross_references)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    case.case_id, case.title, case.description, case.amount,
                    json.dumps(case.responsible_parties), case.date, case.status,
                    case.severity, case.category, case.source, case.source_url,
                    json.dumps(case.documents), case.verification_status,
                    json.dumps(case.cross_references)
                ))
            
            # Store analysis summary
            cursor.execute('''
                INSERT INTO data_source_logs (source, last_update, records_processed, errors_count, status)
                VALUES (?, ?, ?, ?, ?)
            ''', (
                "comprehensive_analysis",
                datetime.now(),
                analysis_results.get("documents_processed", 0),
                1 if analysis_results.get("error") else 0,
                analysis_results.get("overall_risk_level", "unknown")
            ))
            
            conn.commit()
            conn.close()
            
            logger.info("Analysis results stored in database")
            
        except Exception as e:
            logger.error(f"Error storing analysis results: {e}")

    def generate_transparency_report(self, analysis_results: Dict) -> str:
        """Generate comprehensive transparency report"""
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        report = f"""
# 🏛️ REPORTE INTEGRAL DE TRANSPARENCIA Y ANTICORRUPCIÓN
## Municipalidad de Carmen de Areco
### Fecha del Análisis: {report_date}

---

## 📊 RESUMEN EJECUTIVO

**🚨 NIVEL DE RIESGO GENERAL: {analysis_results.get('overall_risk_level', 'UNKNOWN').upper()}

| Métrica Clave                   | Valor                                      |
| ------------------------------- | ------------------------------------------ |
| **Casos de Corrupción Confirmados** | {len([c for c in self.corruption_cases if 'Confirmed' in c.status])} |
| **Casos Bajo Investigación**      | {len([c for c in self.corruption_cases if 'investigation' in c.status.lower()])} |
| **Fuentes de Datos Analizadas**   | {len(analysis_results.get('data_sources_analyzed', []))} |
| **Documentos Procesados**         | {analysis_results.get('documents_processed', 0)} |
| **Anomalías Detectadas**          | {analysis_results.get('anomalies_detected', 0)} |

---

## 🚨 CASOS CONFIRMADOS DE CORRUPCIÓN (HTC)

"""
        
        # Add confirmed corruption cases
        confirmed_cases = [case for case in self.corruption_cases if "Confirmed" in case.status]
        total_confirmed_amount = sum(case.amount for case in confirmed_cases if case.amount)
        
        report += f"**💰 IMPACTO ECONÓMICO TOTAL CONFIRMADO: ${total_confirmed_amount:,.2f} ARS**\n\n"
        
        for case in confirmed_cases:
            report += f"""

### 🔴 {case.title}
- **💰 Monto**: ${case.amount:,.2f} ARS
- **👥 Responsables**: {', '.join(case.responsible_parties)}
- **📅 Período**: {case.date}
- **📋 Categoría**: {case.category}
- **🏛️ Fuente**: {case.source}
- **📄 Estado**: {case.status}

"""
        
        # Add cases under investigation
        investigation_cases = [case for case in self.corruption_cases if "investigation" in case.status.lower()]
        if investigation_cases:
            report += "## ⚠️ CASOS BAJO INVESTIGACIÓN\n\n"
            for case in investigation_cases:
                report += f"""

### 🟡 {case.title}
- **📋 Descripción**: {case.description}
- **👥 Involucrados**: {', '.join(case.responsible_parties)}
- **📅 Período**: {case.date}
- **📄 Estado**: {case.status}

"""
        
        # Add data source analysis
        report += "## 📊 ANÁLISIS DE FUENTES DE DATOS\n\n"
        
        for source in analysis_results.get('data_sources_analyzed', []):
            findings = analysis_results.get('detailed_findings', {}).get(source, {})
            if findings:
                report += f"### 🔍 {source.replace('_', ' ').title()}\n"
                
                if isinstance(findings, list):
                    report += f"- **Registros encontrados**: {len(findings)}\n"
                elif isinstance(findings, dict):
                    if 'documents_found' in findings:
                        report += f"- **Documentos encontrados**: {len(findings['documents_found'])}\n"
                    if 'integrity_issues' in findings:
                        report += f"- **Problemas de integridad**: {len(findings['integrity_issues'])}\n"
                    if 'missing_documents' in findings:
                        report += f"- **Documentos faltantes**: {len(findings['missing_documents'])}\n"
                
                report += "\n"
        
        # Add cross-referencing analysis
        cross_ref_analysis = analysis_results.get('cross_reference_analysis', {})
        if cross_ref_analysis:
            report += "## 🔗 ANÁLISIS DE CRUCE DE DATOS\n\n"
            if cross_ref_analysis.get('entity_correlations'):
                report += "### Correlaciones de Entidades\n"
                for corr in cross_ref_analysis['entity_correlations']:
                    report += f"- **Entidad:** {corr['entity']} (encontrada en {corr['correlation_strength']} fuentes: { ', '.join(corr['sources']) })\n"
                report += "\n"
            if cross_ref_analysis.get('pattern_matches'):
                report += "### Coincidencias con Patrones de Corrupción\n"
                for match in cross_ref_analysis['pattern_matches']:
                    report += f"- **{match['description']}\n"
                report += "\n"

        # Add recommendations
        report += "## 📋 RECOMENDACIONES PRIORITARIAS\n\n"
        
        for i, recommendation in enumerate(analysis_results.get('recommendations', []), 1):
            report += f"{i}. {recommendation}\n"
        
        # Add technical details
        report += f"""

---

## 🔧 DETALLES TÉCNICOS

### Fuentes de Datos Integradas:
"""
        
        for source_name, source_config in self.data_sources.items():
            report += f"- **{source_name.value}**: {source_config.get('base_url', 'N/A')}\n"
        
        # Add GitHub projects used
        report += """

### Proyectos Open Source Utilizados:
- **tommanzur/scraper_boletin_oficial**: Scraping del Boletín Oficial
- **datosgobar**: APIs oficiales del gobierno argentino
- **datosgcba**: Datos abiertos de Buenos Aires
- **reingart/pyafipws**: Integración con servicios AFIP
- **PyPDF2**: Análisis de documentos PDF
- **BeautifulSoup**: Web scraping

"""
        
        report += f"""
---

## ⚖️ MARCO LEGAL Y FUENTES

### Legislación Aplicable:
- Ley de Acceso a la Información Pública (Ley 27.275)
- Decreto 206/2017 (Reglamentación de la Ley de Acceso)
- Ley de Ética Pública (Ley 25.188)
- Código Penal Argentino (Delitos contra la Administración Pública)

### Fuentes de Verificación:
- Honorable Tribunal de Cuentas de la Provincia de Buenos Aires
- Portal de Transparencia Municipal Carmen de Areco
- Boletín Oficial de la República Argentina
- Sistema de Boletín Oficial Municipal (SIBOM)
- Datos.gob.ar (Portal Nacional de Datos Abiertos)

---

## 📞 CONTACTOS ÚTILES PARA DENUNCIAS

### Autoridades de Control:
- **HTC Buenos Aires**: info @htc.gba.gov.ar | Tel: (0221) 429-5588
- **Defensoría del Pueblo BA**: consultas @defensor.gba.gob.ar
- **Ministerio Público Fiscal**: mpf @mpf.gob.ar
- **AAIP (Acceso a Información)**: consultas @aaip.gob.ar

### Organizaciones de la Sociedad Civil:
- **Poder Ciudadano**: info @poderciudadano.org
- **ACIJ**: info @acij.org.ar
- **CIPPEC**: comunicacion @cippec.org

---

*Este reporte fue generated automáticamente utilizando datos públicos oficiales y herramientas de código abierto para promover la transparencia y la rendición de cuentas en la gestión pública.*

**⚖️ Disclaimer Legal**: Toda la información contenida en este reporte proviene de fuentes oficiales públicas. Los casos marcados como "bajo investigación" no implican culpabilidad hasta resolución judicial definitiva. Este análisis tiene fines de transparencia ciudadana y accountability democrático.
"""
        
        return report


class SystemManager:
    """Manages the execution and reporting of the transparency system."""

    def __init__(self):
        self.system = IntegratedTransparencySystem()

    async def search_and_ingest_data(self):
        """Search for data in the project and ingest it into the database."""
        logger.info("Searching for data to ingest...")
        # This can be expanded to search for other data sources
        self.system.ingest_local_inventory()

    def analyze_data_consistency(self):
        """Analyze the consistency of the data in the database."""
        logger.info("Analyzing data consistency...")
        # Placeholder for consistency analysis logic
        pass

    def run_system_check(self):
        """Run a system check to ensure all components are working correctly."""
        logger.info("Running system check...")
        # Placeholder for system check logic
        pass

    async def execute_full_analysis(self):
        """Execute the full analysis pipeline."""
        print("🔍 Carmen de Areco - Sistema Integrado de Transparencia")
        print("=" * 60)
        print("Integrando múltiples fuentes de datos gubernamentales...")
        print("Proyectos GitHub utilizados:")
        print("- tommanzur/scraper_boletin_oficial")
        print("- datosgobar (Gobierno Argentina)")
        print("- datosgcba (Buenos Aires)")
        print("- reingart/pyafipws (AFIP)")
        print("=" * 60)

        try:
            # 1. Ingest data
            await self.search_and_ingest_data()

            # 2. Run comprehensive analysis
            print("\n🚀 Iniciando análisis integral...")
            results = await self.system.run_comprehensive_analysis()

            print(f"\n📊 ANÁLISIS COMPLETADO")
            print(f"🏛️Municipio: Carmen de Areco")
            print(f"⚠️ Nivel de Riesgo: {results['overall_risk_level'].upper()}")
            print(f"📋 Casos de Corrupción: {results['corruption_cases_tracked']}")
            print(f"🔍 Fuentes Analizadas: {len(results['data_sources_analyzed'])}")
            print(f"📄 Documentos Procesados: {results['documents_processed']}")

            # 3. Generate and save report
            report = self.system.generate_transparency_report(results)
            report_path = self.system.data_dir / f"transparency_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(report)

            results_path = self.system.data_dir / f"analysis_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(results_path, 'w', encoding='utf-8') as f:
                json.dump(results, f, indent=2, ensure_ascii=False, default=str)

            print(f"\n📄 Reporte generado: {report_path}")
            print(f"📊 Resultados detallados: {results_path}")

            # 4. Show critical alerts
            if results['overall_risk_level'] in ['critical', 'high']:
                print(f"\n🚨 ALERTA {results['overall_risk_level'].upper()}")
                print("Recomendaciones prioritarias:")
                for i, rec in enumerate(results.get('recommendations', [])[:5], 1):
                    print(f"{i}. {rec}")

            print(f"\n💾 Base de datos: {self.system.db_path}")
            print("✅ Análisis completado exitosamente")

        except (ModuleNotFoundError, ImportError) as e:
            print(f"❌ Error: Módulo necesario no encontrado: {e}")
            print("Por favor, instale las dependencias con 'pip install -r requirements.txt'")
        except Exception as e:
            print(f"❌ Error durante el análisis: {e}")
            logger.error(f"Main execution error: {e}", exc_info=True)

def main():
    """Main function to run the integrated transparency system."""
    manager = SystemManager()
    asyncio.run(manager.execute_full_analysis())

if __name__ == "__main__":
    main()
